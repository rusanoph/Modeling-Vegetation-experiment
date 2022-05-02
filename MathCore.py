import openpyxl  # Работа с Excel.
import scipy.integrate as integrate  # Итегрирование.
from scipy import stats  # Для расчета квантилей распределения Стюдента
from math import sqrt, log, exp, pi  # Элементарные функции и пи 3.141592656358979...

ROOT_SYSTEM = 0
GREEN_PART = 1
BOTH_PART = 2
""" Флаги
flag_ab - определяет конец столбца таблицы
flag_12 - определяет конец строки таблицы
"""


def get_probit_cell(metal: str, anion: str, culture: str, part: int):
	"""
	Вход: Название металла, аниона, культуры и части растения.
	Выход: возвращает список двух элементов вида ['A1', 'B1'], которые.
	задают пробит-функцию для коревой системы и зелёной части соответственно.
	"""
	# Проверка корректности введёной части растения.
	if part != GREEN_PART and part != ROOT_SYSTEM:
		return f"Неверный параметр part:{part}. Введите part равый одной из констант: GREEN_PART, ROOT_SYSTEM."
	# Открываем бд, лист пробит-функций
	workbook = openpyxl.reader.excel.load_workbook(filename="db_toxic.xlsx")
	workbook.active = 1
	active_sheet = workbook.active

	# Приведение к универсальнову виду, чтобы не было различий между,
	# например, Zn, ZN, zn и zN.
	metal = metal.lower()
	anion = anion.lower()
	culture = culture.lower()

	flag_ab = False  # Флаг для проверки, что первый цикл отработал или нет.
	flag_12 = False  # Аналогичный флаг, но для второго цикла.
	# Нужны для определения, пусты или нет проверяемые ячейки.

	metal_row = 3  # Определяет y координату по металлу ячейки пробит-функции.
	anion_row = 3  # Определяет y координату по аниону ячейки пробит-функции.
	while active_sheet['B'+str(metal_row)].value is not None:
		# Блок try-except нужен для избежания ошибки, о невозможности
		# применения метода .lower() к None.
		try:
			if active_sheet['A'+str(metal_row)].value.lower() == metal:
				if active_sheet['B'+str(anion_row)].value.lower() == anion:
					flag_ab = True
					break
				else:
					anion_row += 1
			else:
				metal_row += 1
				anion_row += 1
		except AttributeError:
			metal_row += 1
			anion_row += 1

	ch = 'C'  # Определяет x координату по культуре ячейки пробит-функцции.
	while active_sheet[ch+'1'].value is not None:
		if active_sheet[ch+'1'].value.lower() == culture.lower():
			flag_12 = True
			break
		else:
			ch = chr(ord(ch) + 2)

	if not flag_ab:
		return f"Введённый металл или анион({metal} или {anion}) отсутствует в базе данных."
	if not flag_12:
		return f"Введённая культура({culture}) отсутствует в базе данных."

	return [ch+str(anion_row), chr(ord(ch) + 1)+str(anion_row)][part]


def get_coefficient(probit_f: str):
	"""
	Вход: Пробит-функция в виде строки. Пример: "y = 1,525x + 3,119".
	Выход: Список двух элементов - параметры a(наклон) и b(свободный член).
	Пример: (1,525, 3,119).

	* Допускается смешанное использование разделителей, например, "y = 1.525x + 3,119",
	где присутстует как точка, так и запятая.
	"""
	# Список всех цифр.
	num_chars = [str(chr(n)) for n in range(ord('0'), ord('9') + 1)]
	ab = ""  # На выходе содежит в себе коэфициенты в том порядке,
	# в котором находятся в БД.

	for char in range(0, len(probit_f)):
		# Если символ цифра, точка или запятая, записывается этот символ. Иначе, пробел.
		if probit_f[char] in num_chars or probit_f[char] == ',' or probit_f[char] == '.':
			ab += probit_f[char]
		else:
			ab += ' '

	# Замена символов ',' на '.' и дальнейшее преобразование строк в числа.
	return [float(x) for x in ".".join(ab.split(",")).split()]


def get_ingib(coefficients: list, concentrate: float):
	"""
	Вход: Коэфициенты пробит-функции. Концентрация соединения в мг/л.
	Выход: Процент ингибирования
	"""
	x = log(concentrate, 10)
	y = coefficients[0] * x + coefficients[1]

	# Расчитывается, как функция Лапласа F(y - 5)
	ingib = (1/sqrt(2*pi)) * integrate.romberg(lambda t: exp(-t*t/2), -10, y-5)

	return ingib


def dispersion(n: int, culture: str, part: int):
	"""
	Вход: n - размер выборки; culture - культура растения; part - часть растения.
	Выход: исправленная дисперсия выборки.
	"""
	# Проверка корректности введёной части растения.
	if part != GREEN_PART and part != ROOT_SYSTEM:
		return "Неверный параметр part:{part}. Введите part равый одной из констант: GREEN_PART, ROOT_SYSTEM."
	# Открываем бд, лист выборки
	workbook = openpyxl.reader.excel.load_workbook(filename="db_toxic.xlsx")
	workbook.active = 0
	active_sheet = workbook.active

	flag_12 = False

	ch = 'B'  # Определяет x координату по культуре ячейки выборки.
	while active_sheet[ch + '1'].value is not None:
		if active_sheet[ch + '1'].value.lower() == culture.lower():
			flag_12 = True
			break
		else:
			ch = chr(ord(ch) + 2)

	if not flag_12:
		return f"Введённая культура({culture}) отсутствует в базе данных."

	# Так как в бд идёт сначала вершки и потом корешки
	if part == ROOT_SYSTEM:
		ch = chr(ord(ch) + 1)

	# Вычисление среднего по выборке из n
	mean = 0
	for i in range(1, n + 1):
		mean += active_sheet[ch+str(i + 2)].value
	mean = mean / n

	# Вычисление исправленной дисперсии (СКО)
	disp = 0
	for i in range(1, n + 1):
		disp += pow(active_sheet[ch+str(i + 2)].value - mean, 2)
	disp = sqrt(disp / (n - 1))

	return disp


def get_control_mm(culture: str, part: int):
	"""
	Вход: culture - культура растения; part - часть растения(GREEN_PART или ROOT_SYSTEM).
	Выход: Контрольное значение размера растения в мм.
	"""
	# Проверка корректности введёной части растения.
	if part != GREEN_PART and part != ROOT_SYSTEM:
		return "Неверный параметр part:{part}. Введите part равый одной из констант: GREEN_PART, ROOT_SYSTEM."
	# Открываем бд, лист пробит-функций
	workbook = openpyxl.reader.excel.load_workbook(filename="db_toxic.xlsx")
	workbook.active = 1
	active_sheet = workbook.active

	flag_12 = False

	ch = 'C'  # Определяет x координату по культуре ячейки пробит-функцции.
	while active_sheet[ch + '1'].value is not None:
		if active_sheet[ch + '1'].value.lower() == culture.lower():
			flag_12 = True
			break
		else:
			ch = chr(ord(ch) + 2)

	# Так как в бд идёт сначала вершки и потом корешки
	if part == GREEN_PART:
		ch = chr(ord(ch) + 1)

	# Список всех цифр.
	num_chars = [str(chr(n)) for n in range(ord('0'), ord('9') + 1)]
	ab = ""  # На выходе содержит контрольный размер части растения.
	control_str = active_sheet[ch+'2'].value
	for char in range(0, len(control_str)):
		# Если символ цифра, точка или запятая, записывается этот символ. Иначе, пробел.
		if control_str[char] in num_chars or control_str[char] == ',' or control_str[char] == '.':
			ab += control_str[char]
		else:
			ab += ' '
	if not flag_12:
		return f"Введённая культура({culture}) отсутствует в базе данных."

	return float("".join(ab.split()))


def get_confidence_interval(n: int, culture: str, part: int, alpha: float):
	"""
	Вход: n - размер выборки; culture - культура растения; part - часть растения;
	alpha - уровень значимости
	Выход: границы доверительного интервала
	"""
	ghamma = 1 - alpha
	quantile = stats.t.ppf(ghamma, n - 1)
	return dispersion(n, culture, part) * quantile / sqrt(n)


# Код для проверки математического ядра.
# Пример работы с фукнциями.

# metal_in = input("Введите катион (металл): ")
# anion_in = input("Введите анион: ")
# culture_in = input("Введите культуру: ")
# concentrate_in = float(input("Введите концентрацию"))
#
# wb = openpyxl.reader.excel.load_workbook(filename="db_toxic.xlsx")
# wb.active = 1
# sheet = wb.active
#
#
# y = sheet[get_probit_cell(metal_in, anion_in, culture_in, GREEN_PART)].value
#
# print(get_coefficient(y))
#
#
# print(y)
# print(get_ingib(get_coefficient(y), concentrate_in))
# print(dispersion(10, culture_in, GREEN_PART))
#
# control = get_control_mm(culture_in, GREEN_PART)
# ingib = get_ingib(get_coefficient(y), concentrate_in)
#
# print(f"\n\nИтого: {control*(1 - ingib)} "
# f"+- {get_confidence_interval(30, culture_in, GREEN_PART, 0.1)}")
